# fitz_ai/tabular/direct_query.py
"""
Direct table query - fast path for querying tables without RAG pipeline.

Flow:
1. Read headers only from file (fast)
2. LLM selects relevant columns for the query
3. Parse only those columns from file
4. Create/update ephemeral PostgreSQL table
5. LLM generates SQL
6. Execute and return results
7. Cleanup (optional)

This bypasses embedding entirely, giving sub-second query responses.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING, Any

from fitz_ai.llm.factory import ChatFactory, ModelTier
from fitz_ai.logging.logger import get_logger
from fitz_ai.tabular.store.postgres import (
    PostgresTableStore,
    _sanitize_column_name,
)

if TYPE_CHECKING:
    pass

logger = get_logger(__name__)

# Supported table file extensions
TABLE_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xls"}


@dataclass
class DirectQueryResult:
    """Result of a direct table query."""

    sql: str
    columns: list[str]
    rows: list[list[Any]]
    row_count: int
    answer: str
    table_id: str
    columns_used: list[str]


def is_table_file(path: Path) -> bool:
    """Check if file is a supported table format."""
    return path.suffix.lower() in TABLE_EXTENSIONS


def compute_file_hash(file_path: Path) -> str:
    """Compute hash of file for change detection."""
    hasher = hashlib.md5()
    with open(file_path, "rb") as f:
        # Read in chunks for large files
        for chunk in iter(lambda: f.read(65536), b""):
            hasher.update(chunk)
    return hasher.hexdigest()[:16]


def read_headers(file_path: Path) -> list[str]:
    """
    Read only the headers from a table file (fast).

    Args:
        file_path: Path to CSV/TSV/Excel file.

    Returns:
        List of column headers.
    """
    suffix = file_path.suffix.lower()

    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            headers = next(reader, [])
        return headers

    elif suffix in {".xlsx", ".xls"}:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [
                str(cell.value) if cell.value else f"col_{i}"
                for i, cell in enumerate(next(ws.iter_rows(max_row=1)))
            ]
            wb.close()
            return headers
        except ImportError:
            raise ImportError("openpyxl required for Excel files: pip install openpyxl")

    raise ValueError(f"Unsupported file type: {suffix}")


def parse_columns(
    file_path: Path,
    columns: list[str],
    all_headers: list[str] | None = None,
) -> tuple[list[str], list[list[str]]]:
    """
    Parse only specific columns from a table file.

    Args:
        file_path: Path to table file.
        columns: Column names to extract.
        all_headers: All headers (if already read, avoids re-reading).

    Returns:
        Tuple of (headers, rows) for selected columns only.
    """
    suffix = file_path.suffix.lower()

    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            headers = next(reader, [])

            # Find column indices
            col_indices = []
            for col in columns:
                try:
                    col_indices.append(headers.index(col))
                except ValueError:
                    logger.warning(f"Column '{col}' not found in headers")

            if not col_indices:
                return [], []

            # Extract only needed columns
            selected_headers = [headers[i] for i in col_indices]
            rows = []
            for row in reader:
                selected_row = []
                for i in col_indices:
                    if i < len(row):
                        selected_row.append(row[i])
                    else:
                        selected_row.append("")
                rows.append(selected_row)

        return selected_headers, rows

    elif suffix in {".xlsx", ".xls"}:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            # Get headers
            header_row = next(ws.iter_rows(max_row=1))
            headers = [
                str(cell.value) if cell.value else f"col_{i}" for i, cell in enumerate(header_row)
            ]

            # Find column indices
            col_indices = []
            for col in columns:
                try:
                    col_indices.append(headers.index(col))
                except ValueError:
                    logger.warning(f"Column '{col}' not found in headers")

            if not col_indices:
                wb.close()
                return [], []

            # Extract only needed columns
            selected_headers = [headers[i] for i in col_indices]
            rows = []
            for row in ws.iter_rows(min_row=2):
                selected_row = []
                for i in col_indices:
                    if i < len(row):
                        val = row[i].value
                        selected_row.append(str(val) if val is not None else "")
                    else:
                        selected_row.append("")
                rows.append(selected_row)

            wb.close()
            return selected_headers, rows

        except ImportError:
            raise ImportError("openpyxl required for Excel files: pip install openpyxl")

    raise ValueError(f"Unsupported file type: {suffix}")


def read_sample_values(file_path: Path, num_rows: int = 20) -> dict[str, list[str]]:
    """
    Read sample values for each column (fast - just first few rows).

    Args:
        file_path: Path to table file.
        num_rows: Number of sample rows to read.

    Returns:
        Dict mapping column name to list of unique sample values.
    """
    suffix = file_path.suffix.lower()

    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            headers = next(reader, [])

            # Collect values per column
            col_values: dict[str, set[str]] = {h: set() for h in headers}
            for i, row in enumerate(reader):
                if i >= num_rows:
                    break
                for j, val in enumerate(row):
                    if j < len(headers) and val:
                        col_values[headers[j]].add(val)

        return {h: list(v)[:3] for h, v in col_values.items()}  # Max 3 unique values

    elif suffix in {".xlsx", ".xls"}:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            header_row = next(ws.iter_rows(max_row=1))
            headers = [
                str(cell.value) if cell.value else f"col_{i}" for i, cell in enumerate(header_row)
            ]

            col_values: dict[str, set[str]] = {h: set() for h in headers}
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=num_rows + 1)):
                for j, cell in enumerate(row):
                    if j < len(headers) and cell.value:
                        col_values[headers[j]].add(str(cell.value))

            wb.close()
            return {h: list(v)[:3] for h, v in col_values.items()}

        except ImportError:
            return {}

    return {}


@dataclass
class DirectTableQuery:
    """
    Direct table query engine - fast path without RAG pipeline.

    Usage:
        from fitz_ai.llm import get_chat_factory

        factory = get_chat_factory("cohere")
        query = DirectTableQuery(chat_factory=factory)
        result = query.query(Path("sales.csv"), "how many cars sold in 2005?")
        print(result.answer)
    """

    chat_factory: ChatFactory
    collection: str = "_direct_tables"
    max_results: int = 100
    _table_store: PostgresTableStore | None = field(default=None, repr=False)

    # Tier assignments per task (developer decision)
    TIER_COLUMN_SELECT: ModelTier = "fast"
    TIER_SQL_GENERATE: ModelTier = "balanced"
    TIER_ANSWER_GENERATE: ModelTier = "fast"

    COLUMN_SELECT_PROMPT = """Select columns needed to answer this question.

Columns with sample values:
{columns_with_samples}

Question: {question}

Include columns for filtering (WHERE), grouping (GROUP BY), and results.
Return ONLY a JSON array: ["col1", "col2"]"""

    SQL_PROMPT = """Write a PostgreSQL SQL query to answer this question.

Question: {question}

Table name: {table_name}
Columns: {columns}
Sample data: {samples}

Examples:
- "how many X in 2005?" -> SELECT SUM(CAST("units_sold" AS NUMERIC)) FROM "table" WHERE "year_sold" = '2005'
- "total sales?" -> SELECT SUM(CAST("amount" AS NUMERIC)) FROM "table"
- "which car?" -> SELECT "car_brand", "model" FROM "table" WHERE ...

Rules:
1. Use double quotes around column names: "column_name"
2. For sums/counts of numbers: CAST("column" AS NUMERIC)
3. Text values are exact strings: WHERE "year_sold" = '2005' (not LIKE)
4. Table name is: {table_name}

Write ONLY the SQL query:"""

    ANSWER_PROMPT = """Based on this SQL query result, answer the user's question concisely.

Question: {question}
SQL: {sql}
Results:
{results}

Provide a clear, direct answer based on the data. If the results are empty, say so."""

    def __post_init__(self):
        if self._table_store is None:
            self._table_store = PostgresTableStore(self.collection)

    @property
    def table_store(self) -> PostgresTableStore:
        return self._table_store

    def query(
        self,
        file_path: Path,
        question: str,
        persist: bool = False,
    ) -> DirectQueryResult:
        """
        Query a table file directly.

        Args:
            file_path: Path to table file (CSV/TSV/Excel).
            question: Natural language question.
            persist: If True, keep table for future queries.

        Returns:
            DirectQueryResult with SQL, results, and answer.
        """
        file_path = Path(file_path).resolve()

        if not file_path.exists():
            raise FileNotFoundError(f"Table file not found: {file_path}")

        if not is_table_file(file_path):
            raise ValueError(f"Not a supported table file: {file_path.suffix}")

        # Generate stable table_id from file path
        table_id = hashlib.md5(str(file_path).encode()).hexdigest()[:12]
        current_file_hash = compute_file_hash(file_path)

        try:
            # Step 1: Read headers and sample values (fast - just first few rows)
            all_headers = read_headers(file_path)
            samples = read_sample_values(file_path)
            logger.debug(f"Read {len(all_headers)} headers from {file_path.name}")

            # Step 2: LLM selects relevant columns (with sample context)
            needed_columns = self._select_columns(question, all_headers, samples)
            logger.debug(f"LLM selected columns: {needed_columns}")

            # Step 3: Check if table exists and has needed columns
            stored_file_hash = self.table_store.get_file_hash(table_id)
            table_exists = self.table_store.get_table_name(table_id) is not None

            if table_exists and stored_file_hash and stored_file_hash == current_file_hash:
                # Table exists and file unchanged - check which columns we already have
                existing, missing = self.table_store.has_columns(table_id, needed_columns)
                logger.debug(f"Existing columns: {existing}, Missing: {missing}")

                if missing:
                    # Parse and add missing columns
                    headers, rows = parse_columns(file_path, missing, all_headers)
                    if headers:
                        # Transpose: column_values[row][col]
                        column_values = rows
                        self.table_store.add_columns(table_id, headers, column_values)
                        logger.debug(f"Added {len(missing)} columns incrementally")
            else:
                # Table doesn't exist, file changed, or new - parse needed columns and store fresh
                logger.debug("Creating fresh table (new or changed)")
                headers, rows = parse_columns(file_path, needed_columns, all_headers)

                if not headers:
                    raise ValueError(f"Could not parse columns: {needed_columns}")

                self.table_store.store(
                    table_id=table_id,
                    columns=headers,
                    rows=rows,
                    source_file=str(file_path),
                    file_hash=current_file_hash,
                )

            # Step 4: Get table name and generate SQL
            pg_table_name = self.table_store.get_table_name(table_id)
            columns_info = self.table_store.get_columns(table_id)

            if not pg_table_name or not columns_info:
                raise ValueError(f"Failed to access table {table_id}")

            sanitized_cols, original_cols = columns_info

            # Get sample data for SQL generation
            sample_rows = self._get_sample_data(pg_table_name, sanitized_cols)

            # Step 5: Generate SQL
            sql = self._generate_sql(question, pg_table_name, sanitized_cols, sample_rows)
            logger.debug(f"Generated SQL: {sql}")

            # Step 6: Execute (with retry logic for missing columns)
            result = self.table_store.execute_query(table_id, sql)

            if result is None:
                # Check if SQL references columns not in the table
                missing_cols = self._find_missing_columns(sql, original_cols, all_headers)

                if missing_cols:
                    # Add missing columns and retry
                    logger.debug(f"Adding missing columns for retry: {missing_cols}")
                    headers, rows = parse_columns(file_path, missing_cols, all_headers)
                    if headers:
                        self.table_store.add_columns(table_id, headers, rows)
                        needed_columns.extend(missing_cols)

                        # Refresh column info
                        columns_info = self.table_store.get_columns(table_id)
                        if columns_info:
                            sanitized_cols, original_cols = columns_info
                            sample_rows = self._get_sample_data(pg_table_name, sanitized_cols)

                # Regenerate SQL with updated columns
                sql = self._generate_sql(
                    question,
                    pg_table_name,
                    sanitized_cols,
                    sample_rows,
                    previous_error="Query failed. Use only these columns: "
                    + ", ".join(sanitized_cols),
                )
                result = self.table_store.execute_query(table_id, sql)

            if result is None:
                raise ValueError("SQL query failed after retry")

            col_names, rows = result

            # Step 7: Generate answer
            answer = self._generate_answer(question, sql, col_names, rows)

            # Step 8: Cleanup if not persisting
            if not persist:
                self.table_store.delete(table_id)

            return DirectQueryResult(
                sql=sql,
                columns=col_names,
                rows=rows,
                row_count=len(rows),
                answer=answer,
                table_id=table_id,
                columns_used=needed_columns,
            )

        except Exception:
            # Cleanup on error
            if not persist:
                try:
                    self.table_store.delete(table_id)
                except Exception:
                    pass
            raise

    def _select_columns(
        self, question: str, columns: list[str], samples: dict[str, list[str]]
    ) -> list[str]:
        """Use LLM to select relevant columns."""
        # Format columns with sample values
        lines = []
        for col in columns:
            vals = samples.get(col, [])
            if vals:
                lines.append(f"- {col}: {', '.join(vals)}")
            else:
                lines.append(f"- {col}")

        prompt = self.COLUMN_SELECT_PROMPT.format(
            columns_with_samples="\n".join(lines),
            question=question,
        )
        chat = self.chat_factory(self.TIER_COLUMN_SELECT)
        response = chat.chat([{"role": "user", "content": prompt}])
        return self._parse_json_list(response, fallback=columns[:5])

    def _generate_sql(
        self,
        question: str,
        table_name: str,
        columns: list[str],
        sample_rows: list[list[str]],
        previous_error: str | None = None,
    ) -> str:
        """Generate SQL query."""
        samples: dict[str, list[str]] = {}
        for i, col in enumerate(columns):
            samples[col] = [row[i] for row in sample_rows if i < len(row)]

        error_context = ""
        if previous_error:
            error_context = f"\n\nPrevious error: {previous_error}\nPlease fix the query."

        prompt = (
            self.SQL_PROMPT.format(
                table_name=table_name,
                columns=columns,
                samples=samples,
                question=question,
                max_results=self.max_results,
            )
            + error_context
        )

        chat = self.chat_factory(self.TIER_SQL_GENERATE)
        response = chat.chat([{"role": "user", "content": prompt}])
        return self._extract_sql(response)

    def _generate_answer(
        self,
        question: str,
        sql: str,
        columns: list[str],
        rows: list[list[Any]],
    ) -> str:
        """Generate natural language answer from results."""
        # Format results as markdown table
        if not rows:
            results_str = "(no results)"
        else:
            lines = []
            lines.append("| " + " | ".join(str(c) for c in columns) + " |")
            lines.append("| " + " | ".join("---" for _ in columns) + " |")
            for row in rows[:20]:  # Limit for LLM context
                cells = [str(v)[:50] if v is not None else "" for v in row]
                lines.append("| " + " | ".join(cells) + " |")
            if len(rows) > 20:
                lines.append(f"... and {len(rows) - 20} more rows")
            results_str = "\n".join(lines)

        prompt = self.ANSWER_PROMPT.format(
            question=question,
            sql=sql,
            results=results_str,
        )

        chat = self.chat_factory(self.TIER_ANSWER_GENERATE)
        return chat.chat([{"role": "user", "content": prompt}])

    def _get_sample_data(
        self, pg_table_name: str, columns: list[str], limit: int = 3
    ) -> list[list[str]]:
        """Fetch sample data from PostgreSQL table."""
        cols_str = ", ".join(f'"{c}"' for c in columns)
        sql = f'SELECT {cols_str} FROM "{pg_table_name}" ORDER BY _row_num LIMIT {limit}'

        result = self.table_store.execute_query("", sql)
        if result:
            _, rows = result
            return [[str(v) if v is not None else "" for v in row] for row in rows]
        return []

    def _parse_json_list(self, response: str, fallback: list[str]) -> list[str]:
        """Parse JSON list from LLM response."""
        import re

        text = response.strip()

        # Remove markdown code blocks
        if text.startswith("```"):
            parts = text.split("```")
            if len(parts) >= 2:
                text = parts[1].strip()
                if text.startswith("json"):
                    text = text[4:].strip()

        # Try direct JSON parse
        try:
            result = json.loads(text)
            if isinstance(result, list):
                # Flatten: if items look like string-repr lists, parse them
                flat = []
                for item in result:
                    if isinstance(item, str) and item.startswith("["):
                        try:
                            inner = json.loads(item.replace("'", '"'))
                            if isinstance(inner, list):
                                flat.extend(str(i) for i in inner)
                                continue
                        except (json.JSONDecodeError, ValueError):
                            pass
                    flat.append(str(item))
                # Filter to only valid column names
                valid = [c for c in flat if c in fallback]
                if valid:
                    return valid
        except (json.JSONDecodeError, ValueError):
            pass

        # Try to find JSON array in text
        match = re.search(r"\[([^\]]+)\]", text)
        if match:
            try:
                arr_text = "[" + match.group(1) + "]"
                # Normalize quotes
                arr_text = arr_text.replace("'", '"')
                result = json.loads(arr_text)
                if isinstance(result, list):
                    valid = [str(c) for c in result if str(c) in fallback]
                    if valid:
                        return valid
            except (json.JSONDecodeError, ValueError):
                pass

        # Extract quoted strings that are valid column names
        matches = re.findall(r'["\']([^"\']+)["\']', response)
        valid = [m for m in matches if m in fallback]
        if valid:
            return valid

        logger.warning("Failed to parse column selection, using fallback")
        return fallback

    def _extract_sql(self, response: str) -> str:
        """Extract SQL from LLM response."""
        text = response.strip()

        if "```" in text:
            match = re.search(r"```(?:sql)?\s*(.*?)```", text, re.DOTALL | re.IGNORECASE)
            if match:
                text = match.group(1).strip()
            else:
                text = text.replace("```sql", "").replace("```", "").strip()

        if not text.upper().startswith("SELECT"):
            match = re.search(r"(SELECT\s+.+)", text, re.DOTALL | re.IGNORECASE)
            if match:
                text = match.group(1)

        return text

    def _find_missing_columns(
        self,
        sql: str,
        current_columns: list[str],
        all_headers: list[str],
    ) -> list[str]:
        """
        Find columns referenced in SQL that aren't in the table.

        Args:
            sql: The SQL query.
            current_columns: Original column names currently in the table.
            all_headers: All available column names from the file.

        Returns:
            List of missing column names that can be added.
        """
        # Extract potential column references from SQL
        # Look for quoted identifiers and unquoted words
        sql_lower = sql.lower()

        # Create lowercase lookup sets
        current_set = {c.lower() for c in current_columns}

        missing = []
        for header in all_headers:
            header_lower = header.lower()
            sanitized = _sanitize_column_name(header)

            # Check if this header is referenced in SQL but not in current columns
            if header_lower not in current_set:
                # Check various forms of the column in SQL
                if (
                    header_lower in sql_lower
                    or sanitized in sql_lower
                    or f'"{header_lower}"' in sql_lower
                    or f'"{sanitized}"' in sql_lower
                ):
                    missing.append(header)

        return missing


__all__ = [
    "DirectTableQuery",
    "DirectQueryResult",
    "is_table_file",
    "read_headers",
    "read_sample_values",
    "parse_columns",
    "compute_file_hash",
]
